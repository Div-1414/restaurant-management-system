from rest_framework import viewsets, status
from rest_framework.decorators import api_view, action
from rest_framework.response import Response
from django.contrib.auth import authenticate
from django.db.models import Q, Sum
from django.utils import timezone
import jwt
import qrcode
import io
import base64
from datetime import datetime, timedelta
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync
from decimal import Decimal
from django.conf import settings
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from restaurant_app.utils import send_owner_account_email


from restaurant_app.utils import ensure_restaurant_is_active

from rest_framework.exceptions import PermissionDenied
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.db.models import Q
from rest_framework.permissions import AllowAny
from rest_framework.decorators import permission_classes
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync


from .models import (
    User, Restaurant, Table, MenuCategory, MenuItem,
    OrderSession, Order, OrderItem, Bill,Hall
)
from .serializers import (
    UserSerializer, UserCreateSerializer, RestaurantSerializer, TableSerializer,
    MenuCategorySerializer, MenuItemSerializer, OrderSerializer, OrderItemSerializer,
    OrderSessionSerializer, BillSerializer
)
from .models import TableGroup


from openpyxl import Workbook
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from restaurant_app.utils import send_restaurant_status_email



# ---------------- AUTH ----------------

@api_view(['POST'])
@permission_classes([AllowAny])
def login(request):
    username = request.data.get('username')
    password = request.data.get('password')

    user = authenticate(username=username, password=password)
    if user:
        token = jwt.encode({
            'user_id': user.id,
            'username': user.username,
            'role': user.role,
            'restaurant_id': str(user.restaurant.id) if user.restaurant else None,
            'exp': datetime.utcnow() + timedelta(days=7)
        }, settings.JWT_SECRET_KEY, algorithm='HS256')

        return Response({
            'token': token,
            'user': UserSerializer(user).data
        })

    return Response({'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)


@api_view(['POST'])
@permission_classes([AllowAny])
def register(request):
    serializer = UserCreateSerializer(data=request.data)
    if serializer.is_valid():
        user = serializer.save()
        return Response(UserSerializer(user).data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# ---------------- RESTAURANTS ----------------

class RestaurantViewSet(viewsets.ModelViewSet):
    queryset = Restaurant.objects.all()
    serializer_class = RestaurantSerializer

     # ✅ Accept BOTH JSON and multipart
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    def create(self, request, *args, **kwargs):
        """
        Super Admin MUST create restaurant with owner credentials
        Required:
        - owner_username
        - owner_password
        Optional:
        - owner_email
        - owner_phone
        - owner_full_name
        """

        data = request.data.copy()

        owner_username = data.get('owner_username')
        owner_password = data.get('owner_password')
        owner_email = data.get('owner_email', '')
        owner_phone = data.get('owner_phone', '')
        owner_full_name = data.get('owner_full_name', '')

        # ✅ Remove owner-only fields from restaurant payload
        data.pop('owner_username', None)
        data.pop('owner_password', None)
        data.pop('owner_email', None)
        data.pop('owner_phone', None)
        data.pop('owner_full_name', None)

        # ❌ BLOCK creation if owner creds missing
        if not owner_username or not owner_password:
            return Response(
                {'error': 'Owner username and password are required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # ❌ Prevent duplicate usernames
        if User.objects.filter(username=owner_username).exists():
            return Response(
                {'error': 'Owner username already exists'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 1️⃣ Create restaurant
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        restaurant = serializer.save()

        try:
            # 2️⃣ Create owner user
            owner = User.objects.create_user(
                username=owner_username,
                password=owner_password,
                email=owner_email,
                role='owner',
                phone=owner_phone,
                restaurant=restaurant,
                first_name=owner_full_name  # full name stored here
            )

            # 3️⃣ Link restaurant ↔ owner
            restaurant.owner = owner
            restaurant.save()
            try:
                send_owner_account_email(
                owner_name=owner_full_name or owner_username,
                restaurant_name=restaurant.name,
                owner_email=owner_email,
                password=owner_password,
                
             )

            except Exception as e:
              print("Email sending failed:", str(e))    

            # ✅ 4️⃣ CREATE DEFAULT HALL (CRITICAL FOR NEW FEATURE)
            Hall.objects.get_or_create(
                restaurant=restaurant,
                name='General'
            )

        except Exception:
            # 🔥 rollback if anything fails
            restaurant.delete()
            return Response(
                {'error': 'Failed to create owner account'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        return Response(
            RestaurantSerializer(restaurant).data,
            status=status.HTTP_201_CREATED
        )

    # ---------------- OWNER OPEN / CLOSE ----------------

    @action(detail=True, methods=['patch'], url_path='toggle-open')
    def toggle_open(self, request, pk=None):
        restaurant = self.get_object()
        restaurant.is_open = not restaurant.is_open
        restaurant.save()

        return Response({
            'id': str(restaurant.id),
            'is_open': restaurant.is_open,
            'message': f"Restaurant is now {'open' if restaurant.is_open else 'closed'}"
        })

    # ---------------- SUPER ADMIN ACTIVE / INACTIVE ----------------

    @action(detail=True, methods=['patch'], url_path='set-status')
    def set_status(self, request, pk=None):
        restaurant = self.get_object()
        raw_status = request.data.get('status')

    # 🔥 BACKWARD + FRONTEND SAFE
        if isinstance(raw_status, dict):
         new_status = raw_status.get('status')
        else:
         new_status = raw_status

        print("DEBUG request.data:", request.data)
        print("DEBUG content-type:", request.content_type)

        if new_status not in ['active', 'inactive']:
            return Response(
                {'error': 'Invalid status value'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        previous_status = restaurant.status
        restaurant.status = new_status
        restaurant.save()

        if previous_status != new_status:

              if restaurant.owner and restaurant.owner.email:
                  try:
                      send_restaurant_status_email( owner_name=restaurant.owner.first_name or restaurant.owner.username, restaurant_name=restaurant.name, owner_email=restaurant.owner.email, status=new_status )
                  except Exception as e: 
                      print("Failed to send status email:", str(e))  
        



        return Response(RestaurantSerializer(restaurant).data)

    # ---------------- STATS ----------------

    @action(detail=True, methods=['get'])
    def stats(self, request, pk=None):
        restaurant = self.get_object()

        # 🔒 BLOCK OWNER IF RESTAURANT INACTIVE
        ensure_restaurant_is_active(restaurant)

        total_tables = restaurant.tables.count()
        occupied_tables = restaurant.tables.filter(status='occupied').count()

        active_sessions = OrderSession.objects.filter(
            table__restaurant=restaurant,
            status='active'
        ).count()

        today = timezone.now().date()

        paid_bills = Bill.objects.filter(
            session__table__restaurant=restaurant,
            created_at__date=today,
            payment_status='paid'
        )

        table_revenue = paid_bills.aggregate(total=Sum('total'))['total'] or 0


        parcel_revenue = ParcelBill.objects.filter(
          parcel_order__restaurant=restaurant,
          created_at__date=today,
          payment_status='paid'
        ).aggregate(total=Sum('total'))['total'] or 0

        today_revenue = table_revenue + parcel_revenue

        return Response({
            'total_tables': total_tables,
            'occupied_tables': occupied_tables,
            'available_tables': total_tables - occupied_tables,
            'active_sessions': active_sessions,
            'today_revenue': str(today_revenue),
            'is_open': restaurant.is_open
        })


# ---------------- TABLES ----------------



from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import Table, Hall
from .serializers import TableSerializer
import qrcode, io, base64


class TableViewSet(viewsets.ModelViewSet):
    queryset = Table.objects.all()
    serializer_class = TableSerializer

    # ================= GET QUERYSET =================
    def get_queryset(self):
      user = self.request.user
      qs = Table.objects.all()

      # 🔒 OWNER RULES
      if hasattr(user, 'role') and user.role == 'owner':
        if not user.restaurant:
            return Table.objects.none()

        qs = qs.filter(restaurant=user.restaurant)

      # ✅ APPLY FILTERS ONLY FOR LIST VIEW
      if self.action == 'list':
        restaurant_id = self.request.query_params.get('restaurant_id')
        if restaurant_id:
            qs = qs.filter(restaurant_id=restaurant_id)

        hall_id = self.request.query_params.get('hall_id')
        if hall_id:
            qs = qs.filter(hall_id=hall_id)

      return qs


    # ================= SINGLE CREATE =================
    def perform_create(self, serializer):
        user = self.request.user

        # 🔒 OWNER
        if hasattr(user, 'role') and user.role == 'owner':
            if not user.restaurant:
                raise PermissionDenied("Owner has no restaurant assigned")

            ensure_restaurant_is_active(user.restaurant)

            # Owner can only create for his restaurant
            serializer.save(restaurant=user.restaurant)
            return

        # 🟢 SUPER ADMIN
        if hasattr(user, 'role') and user.role == 'super_admin':
            serializer.save()
            return

        raise PermissionDenied("You are not allowed to create tables")

    # ================= GENERATE QR =================
    @action(detail=True, methods=['post'])
    def generate_qr(self, request, pk=None):
        user = request.user

        if hasattr(user, 'role') and user.role == 'owner':
            ensure_restaurant_is_active(user.restaurant)

        table = self.get_object()
        frontend_url = request.data.get('frontend_url')

        if not frontend_url:
            return Response(
                {'error': 'frontend_url is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        qr_url = f"{frontend_url}/menu/{table.restaurant.id}/{table.id}"
        table.qr_code = qr_url
        table.save(update_fields=['qr_code'])

        return Response({'qr_url': qr_url})

    # ================= BULK CREATE =================
    @action(detail=False, methods=['post'], url_path='bulk-create')
    def bulk_create(self, request):
        user = request.user

        restaurant_id = (
            request.data.get('restaurant_id')
            or request.data.get('restaurant')
        )

        hall_id = request.data.get('hall_id') or request.data.get('hall')
        frontend_url = request.data.get('frontend_url')

        # SAFE count parsing
        try:
            count = int(request.data.get('table_count') or request.data.get('count'))
        except (TypeError, ValueError):
            count = 0

        if count <= 0 or not frontend_url:
            return Response(
                {'error': 'table_count and frontend_url are required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 🔒 OWNER RULES
        if hasattr(user, 'role') and user.role == 'owner':
            if not user.restaurant:
                raise PermissionDenied("Owner has no restaurant assigned")

            ensure_restaurant_is_active(user.restaurant)
            restaurant_id = str(user.restaurant.id)

        # 🟢 SUPER ADMIN
        elif hasattr(user, 'role') and user.role == 'super_admin':
            if not restaurant_id:
                return Response(
                    {'error': 'restaurant_id is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        else:
            raise PermissionDenied("You are not allowed to create tables")

        # ✅ Resolve hall
        hall = None
        if hall_id:
            hall = Hall.objects.filter(
                id=hall_id,
                restaurant_id=restaurant_id
            ).first()

            if not hall:
                return Response(
                    {'error': 'Invalid hall for this restaurant'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        existing = Table.objects.filter(
            restaurant_id=restaurant_id,
            hall=hall
        ).count()

        tables = [
            Table(
                restaurant_id=restaurant_id,
                hall=hall,
                table_number=existing + i + 1,
                qr_code='TEMP'
            )
            for i in range(count)
        ]

        Table.objects.bulk_create(tables)

        created_tables = Table.objects.filter(
            restaurant_id=restaurant_id,
            hall=hall
        ).order_by('-created_at')[:count]

        for table in created_tables:
            table.qr_code = f"{frontend_url}/menu/{restaurant_id}/{table.id}"
            table.save(update_fields=['qr_code'])

        return Response(
            {
                'message': f'{count} tables created',
                'hall': hall.name if hall else 'General'
            },
            status=status.HTTP_201_CREATED
        )
    #--for single delete---------------#
    def destroy(self, request, *args, **kwargs):
      user = request.user
      table = self.get_object()

      # 🔒 OWNER RULES
      if hasattr(user, 'role') and user.role == 'owner':
        ensure_restaurant_is_active(user.restaurant)

        if table.restaurant != user.restaurant:
           raise PermissionDenied("You cannot delete this table")

            # ❌ NEW RULE: block occupied table delete
        if table.status == 'occupied':
          raise PermissionDenied("You cannot delete an occupied table")

      return super().destroy(request, *args, **kwargs)

    
    @action(detail=False, methods=['delete'], url_path='bulk-delete')
    def bulk_delete(self, request):
      user = request.user

      hall_id = request.data.get('hall_id')

      if not hall_id:
        return Response(
            {'error': 'hall_id is required'},
            status=status.HTTP_400_BAD_REQUEST
        )

     # 🔍 Find hall
      hall = Hall.objects.filter(id=hall_id).first()
      if not hall:
        return Response(
            {'error': 'Hall not found'},
            status=status.HTTP_404_NOT_FOUND
        )

    # 🔒 OWNER RULES
      if hasattr(user, 'role') and user.role == 'owner':
        if not user.restaurant:
            raise PermissionDenied("Owner has no restaurant")

        ensure_restaurant_is_active(user.restaurant)

        if hall.restaurant != user.restaurant:
            raise PermissionDenied("You cannot delete tables of this hall")

    # 🧮 Count tables before delete
      tables_qs = Table.objects.filter(hall=hall)
      tables_count = tables_qs.count()

      if tables_count == 0:
        return Response(
            {'message': 'No tables found in this hall'},
            status=status.HTTP_200_OK
        )

    # ❌ DELETE TABLES
      tables_qs.delete()

      return Response(
        {
            'message': 'Tables deleted successfully',
            'hall': hall.name,
            'tables_deleted': tables_count
        },
        status=status.HTTP_200_OK
    )

     #-------------For combinimg tables--------------#
    @action(detail=False, methods=['post'], url_path='combine')
    def combine_tables(self, request):
      user = request.user
      table_ids = request.data.get('table_ids', [])

      if user.role not in ['owner', 'restaurant_manager']:
        raise PermissionDenied("You are not allowed to combine tables")

      if not table_ids or len(table_ids) < 2:
        return Response(
            {"error": "Select at least 2 tables to combine"},
            status=status.HTTP_400_BAD_REQUEST
        )

      tables = Table.objects.filter(id__in=table_ids)

      if tables.count() != len(table_ids):
        return Response(
            {"error": "One or more tables not found"},
            status=status.HTTP_400_BAD_REQUEST
        )

      restaurant = tables.first().restaurant
      ensure_restaurant_is_active(restaurant)

      # 🔐 Ensure all tables belong to same restaurant
      if any(t.restaurant != restaurant for t in tables):
        return Response(
            {"error": "All tables must belong to same restaurant"},
            status=status.HTTP_400_BAD_REQUEST
        )

      # 🔐 Ensure all tables are available
      for table in tables:
        if table.status != 'available':
            return Response(
                {"error": f"Table {table.table_number} is not available"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Prevent combining table with active session
        active_session = OrderSession.objects.filter(
            table=table,
            status__in=['active', 'bill_generated']
        ).exists()

        if active_session:
            return Response(
                {"error": f"Table {table.table_number} has active session"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Prevent combining if already in active group
        existing_group = table.table_groups.filter(is_active=True).exists()
        if existing_group:
            return Response(
                {"error": f"Table {table.table_number} already in group"},
                status=status.HTTP_400_BAD_REQUEST
            )

      # ✅ Create group
      group = TableGroup.objects.create(
        restaurant=restaurant,
        created_by=user
      )

      group.tables.set(tables)

      # Mark tables occupied
      tables.update(status='occupied')

      return Response(
        {
            "message": "Tables combined successfully",
            "group_id": str(group.id)
        },
        status=status.HTTP_201_CREATED
    )



    #----------For Transfer Table-------------------#
    @action(detail=False, methods=['post'], url_path='transfer')
    def transfer_table(self, request):
       user = request.user
       session_id = request.data.get('session_id')
       target_table_id = request.data.get('target_table_id')

       if user.role not in ['owner', 'restaurant_manager']:
        raise PermissionDenied("You are not allowed to transfer tables")

       try:
        session = OrderSession.objects.get(id=session_id)
       except OrderSession.DoesNotExist:
        return Response({"error": "Session not found"}, status=404)

       # 🔐 Only active sessions can be transferred
       if session.status != 'active':
         return Response(
            {"error": "Only active sessions can be transferred"},
            status=400
        )

       # ❌ Cannot transfer group session
       if session.table_group:
         return Response(
            {"error": "Cannot transfer combined table session"},
            status=400
        )

       try:
         target_table = Table.objects.get(id=target_table_id)
       except Table.DoesNotExist:
         return Response({"error": "Target table not found"}, status=404)

       # 🔐 Same restaurant check
       if session.table.restaurant != target_table.restaurant:
        return Response(
            {"error": "Tables must belong to same restaurant"},
            status=400
        )

       # 🔐 Target must be available
       if target_table.status != 'available':
          return Response(
            {"error": "Target table is not available"},
            status=400
        )

       # 🔐 Target must not be in active group
       if target_table.table_groups.filter (is_active=True).exists():
         return Response(
            {"error": "Target table is part of active group"},
            status=400
        )

       # 🔐 Target must not have active session
       existing_session = OrderSession.objects.filter(
         table=target_table,
         status__in=['active', 'bill_generated']
     ).exists()

       if existing_session:
         return Response(
            {"error": "Target table has active session"},
            status=400
        )

     # 🔁 Perform transfer
       old_table = session.table

       session.table = target_table
       session.save()

       # Update table statuses
       old_table.status = 'available'
       old_table.save()

       target_table.status = 'occupied'
       target_table.save()

       return Response({"message": "Table transferred successfully"})

    #-------for floorview for manager----------------#
    @action(detail=False, methods=['get'], url_path='floor-overview')
    def floor_overview(self, request):
       user = request.user

       if not user.restaurant:
        return Response({"error": "No restaurant"}, status=403)

       restaurant = user.restaurant
       ensure_restaurant_is_active(restaurant)

       tables = Table.objects.filter(
        restaurant=restaurant
    ).select_related("hall").order_by('table_number')

       active_groups = restaurant.table_groups.filter(is_active=True)

       # Build group mapping
       table_group_map = {}

       for group in active_groups:
        group_tables = list(group.tables.values_list('id', flat=True))
        for table_id in group_tables:
            table_group_map[table_id] = {
                "group_id": str(group.id),
                "group_tables": group_tables
            }

       response = []

       for table in tables:

        # 🔥 ACTIVE SESSION
        # 🔍 First check direct table session
        session = OrderSession.objects.filter(
                 table=table,
               status="active"
                       ).first()

# 🔍 If not found, check group session
        if not session:
         group = table.table_groups.filter(is_active=True).first()
         if group:
            session = OrderSession.objects.filter(
            table_group=group,
            status="active"
        ).first()


        has_manager_order = False
        has_qr_order = False

        if session:
            orders = session.orders.all()

            has_manager_order = orders.filter(order_source="manager").exists()
            has_qr_order = orders.filter(order_source="qr").exists()

        group_info = table_group_map.get(table.id)

        response.append({
            "id": str(table.id),
            "table_number": table.table_number,
            "status": table.status,
            "hall_name": table.hall.name if table.hall else "General",

            "is_combined": bool(group_info),
            "group_id": group_info["group_id"] if group_info else None,
            "group_tables": group_info["group_tables"] if group_info else [],

            # 🔥 NEW FLAGS
            "has_manager_order": has_manager_order,
            "has_qr_order": has_qr_order,
        })

       return Response({
        "tables": response,
        "combinedGroups": active_groups.count(),
        "combinedTables": sum(
            group.tables.count() for group in active_groups
        )
        })






# ---------------- MENU ----------------

class MenuCategoryViewSet(viewsets.ModelViewSet):
    queryset = MenuCategory.objects.all()
    serializer_class = MenuCategorySerializer

    def get_queryset(self):
        queryset = MenuCategory.objects.all()
        restaurant_id = self.request.query_params.get('restaurant_id')
        if restaurant_id:
            queryset = queryset.filter(restaurant_id=restaurant_id)
        return queryset


class MenuItemViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = MenuItem.objects.select_related(
        'category',
        'category__restaurant'
    )

    def get_queryset(self):
        queryset = self.queryset
        category_id = self.request.query_params.get('category_id')
        restaurant_id = self.request.query_params.get('restaurant_id')

        if category_id:
            queryset = queryset.filter(category_id=category_id)
        elif restaurant_id:
            queryset = queryset.filter(category__restaurant_id=restaurant_id)

        return queryset

    def get_serializer_class(self):
     return MenuItemSerializer

    # 🔐 Common owner permission check
    def _check_owner_permissions(self, item):
        user = self.request.user

        # Super admin bypass
        if user.role == 'super_admin':
            return

        # Only owner allowed
        if user.role != 'owner':
            raise PermissionDenied("You do not have permission to modify menu items")

        # Restaurant must be active
        ensure_restaurant_is_active(user.restaurant)

        # Item must belong to owner's restaurant
        if item.category.restaurant != user.restaurant:
            raise PermissionDenied("You cannot modify this menu item")

    # 🔁 Toggle availability
    @action(detail=True, methods=['patch'], url_path='toggle-availability')
    def toggle_availability(self, request, pk=None):
        item = self.get_object()
        self._check_owner_permissions(item)

        item.available = not item.available
        item.save(update_fields=['available'])

        return Response(
            {
                "message": "Item availability updated",
                "available": item.available,
            },
            status=status.HTTP_200_OK
        )

    # 🔁 Toggle half option
    @action(detail=True, methods=['patch'], url_path='toggle-half')
    def toggle_half(self, request, pk=None):
        item = self.get_object()
        self._check_owner_permissions(item)

        item.allow_half = not item.allow_half
        if not item.half_price:
             return Response(
           {"detail": "Half price must be set before enabling half plate"},
                status=400
        )

        item.save(update_fields=['allow_half'])

        return Response(
            {
                "message": "Half option updated",
                "allow_half": item.allow_half,
            },
            status=status.HTTP_200_OK
        )

# ---------------- ORDERS ----------------

class OrderViewSet(viewsets.ModelViewSet):
    queryset = Order.objects.all()
    serializer_class = OrderSerializer

    def get_queryset(self):
      user = self.request.user

      queryset = Order.objects.select_related(
        'session__table__restaurant'
      ).prefetch_related(
        'items__menu_item'
      )

      # 🔒 MANAGER: only his active manual orders
      if user.role == 'restaurant_manager':
        queryset = queryset.filter(
            created_by=user,
            order_source='manager',
            session__status='active'
        )

      # 🔒 OWNER: only bill generated sessions
      elif user.role == 'owner':
        queryset = queryset.filter(
            session__table__restaurant=user.restaurant,
            session__status='bill_generated'
        )

      # 🟢 SUPER ADMIN (optional)
      elif user.role == 'super_admin':
        restaurant_id = self.request.query_params.get('restaurant_id')
        if restaurant_id:
            queryset = queryset.filter(
                session__table__restaurant_id=restaurant_id
            )

      return queryset


# ---------------- CREATE ORDER (CUSTOMER) ----------------

@api_view(['POST'])
@permission_classes([AllowAny])
@authentication_classes([])
def create_order(request):
    table_id = request.data.get('table_id')
    items = request.data.get('items', [])
    special_instructions = request.data.get('special_instructions', '')
    customer_name = request.data.get('customer_name', '')
    customer_phone = request.data.get('customer_phone', '')

    try:
        table = Table.objects.get(id=table_id)
        restaurant = table.restaurant

        # 🔒 Restaurant active check
        ensure_restaurant_is_active(restaurant)

        # 🔒 Restaurant open check
        if not restaurant.is_open:
            return Response(
                {'error': 'Restaurant is currently closed', 'is_open': False},
                status=status.HTTP_400_BAD_REQUEST
            )

        # ============================================
        # 🔥 NEW: CHECK IF TABLE BELONGS TO ACTIVE GROUP
        # ============================================

        active_group = table.table_groups.filter(is_active=True).first()

        if active_group:
            # Look for existing session for this group
            session = OrderSession.objects.filter(
                table_group=active_group
            ).exclude(status='paid').first()

            if session:
                if session.status != 'active':
                    return Response(
                        {"error": "Cannot place order. Bill already generated."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            else:
                # Create new session for group
                session = OrderSession.objects.create(
                    table_group=active_group,
                    customer_name=customer_name,
                    customer_phone=customer_phone
                )

        else:
            # ============================================
            # 🔵 NORMAL TABLE FLOW
            # ============================================

            session = OrderSession.objects.filter(
                table=table
            ).exclude(status='paid').first()

            if session:
                if session.status != 'active':
                    return Response(
                        {"error": "Cannot place order. Bill already generated."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            else:
                session = OrderSession.objects.create(
                    table=table,
                    customer_name=customer_name,
                    customer_phone=customer_phone
                )
                table.status = 'occupied'
                table.save()

        # ============================================
        # 🧾 CREATE ORDER
        # ============================================

        order = Order.objects.create(
            session=session,
            special_instructions=special_instructions,
            created_by=request.user if request.user.is_authenticated else None,
            order_source='qr'
        )

        total = 0
        for item_data in items:
            menu_item = MenuItem.objects.get(id=item_data['menu_item_id'])
            quantity = item_data.get('quantity', 1)

            is_half = item_data.get('is_half', False)
            option_ids = item_data.get('option_ids', [])

            # 🔥 Base price
            if is_half and menu_item.allow_half:
                base_price = menu_item.half_price
            else:
                base_price = menu_item.price

            # 🔥 Add option prices
            options = MenuItemOption.objects.filter(
                id__in=option_ids,
                is_active=True
            )

            extra_price = sum(opt.extra_price for opt in options)

            final_unit_price = base_price + extra_price
            final_price = final_unit_price * quantity

            order_item = OrderItem.objects.create(
                order=order,
                menu_item=menu_item,
                quantity=quantity,
                is_half=is_half,
                price=final_price
            )

            order_item.selected_options.set(options)

            total += final_price

        order.total = total
        order.save()

        # ============================================
        # 🔥 NEW: BROADCAST TO KITCHEN DASHBOARD
        # ============================================
        
        # Determine if this is the first order in the session
        all_orders = list(session.orders.all().order_by('created_at'))
        is_first_order = (len(all_orders) == 1) or (all_orders[0].id == order.id)
        
        # Broadcast to kitchen
        broadcast_to_kitchen(order, restaurant.id, is_first_order)

        return Response(
            OrderSerializer(order).data,
            status=status.HTTP_201_CREATED
        )

    except Table.DoesNotExist:
        return Response(
            {'error': 'Table not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    

# ------Manager create order function -------------#

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def manager_create_order(request):
    user = request.user

    if user.role not in ['owner', 'restaurant_manager']:
        raise PermissionDenied("Only owner or manager can place manual orders")

    table_id = request.data.get('table_id')
    items = request.data.get('items', [])
    special_instructions = request.data.get('special_instructions', '')

    try:
        table = Table.objects.get(id=table_id)

        ensure_restaurant_is_active(table.restaurant)

        if table.restaurant != user.restaurant:
            return Response(
                {"error": "Invalid table for your restaurant"},
                status=400
            )

        if not table.restaurant.is_open:
            return Response(
                {"error": "Restaurant is closed"},
                status=400
            )

        # Same session logic as customer
        active_group = table.table_groups.filter(is_active=True).first()

        if active_group:
            session = OrderSession.objects.filter(
                table_group=active_group
            ).exclude(status='paid').first()

            if session:
                if session.status != 'active':
                    return Response(
                        {"error": "Bill already generated"},
                        status=400
                    )
            else:
                session = OrderSession.objects.create(
                    table_group=active_group
                )

        else:
            session = OrderSession.objects.filter(
                table=table
            ).exclude(status='paid').first()

            if session:
                if session.status != 'active':
                    return Response(
                        {"error": "Bill already generated"},
                        status=400
                    )
            else:
                session = OrderSession.objects.create(
                    table=table
                )
                table.status = 'occupied'
                table.save()

        # Create order
        order = Order.objects.create(
            session=session,
            special_instructions=special_instructions,
            created_by=user,
            order_source='manager'
        )

        total = 0
        for item_data in items:
            menu_item = MenuItem.objects.get(id=item_data['menu_item_id'])
            quantity = item_data['quantity']
            
            # 🔥 NEW: Support half plates for manager orders too
            is_half = item_data.get('is_half', False)
            option_ids = item_data.get('option_ids', [])
            
            # Calculate price
            if is_half and menu_item.allow_half:
                base_price = menu_item.half_price
            else:
                base_price = menu_item.price
            
            # Add option prices
            options = MenuItemOption.objects.filter(
                id__in=option_ids,
                is_active=True
            )
            extra_price = sum(opt.extra_price for opt in options)
            
            final_unit_price = base_price + extra_price
            final_price = final_unit_price * quantity

            OrderItem.objects.create(
                order=order,
                menu_item=menu_item,
                quantity=quantity,
                is_half=is_half,
                price=final_price
            )

            total += final_price

        order.total = total
        order.save()

        # ============================================
        # 🔥 NEW: BROADCAST TO KITCHEN DASHBOARD
        # ============================================
        
        # Determine if this is the first order in the session
        all_orders = list(session.orders.all().order_by('created_at'))
        is_first_order = (len(all_orders) == 1) or (all_orders[0].id == order.id)
        
        # Broadcast to kitchen
        broadcast_to_kitchen(order, user.restaurant.id, is_first_order)

        return Response(
            OrderSerializer(order).data,
            status=201
        )

    except Table.DoesNotExist:
        return Response({"error": "Table not found"}, status=404)
    
# ---------------- MANAGER ACTIVE SESSIONS ----------------
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def manager_active_sessions(request):
    user = request.user

    if user.role not in ['owner', 'restaurant_manager']:
        raise PermissionDenied("Not allowed")

    restaurant = user.restaurant

    sessions = OrderSession.objects.filter(
        table__restaurant=restaurant,
        status='active'
    ).prefetch_related(
        'orders__items__menu_item'
    )

    data = []

    for session in sessions:
        # Only sessions where manager created orders
        manager_orders = session.orders.filter(created_by=user)

        if not manager_orders.exists():
            continue

        session_total = sum(order.total for order in manager_orders)

        data.append({
            "id": str(session.id),
            "table_number": session.table.table_number if session.table else "Group",
            "status": session.status,
            "total": session_total,
            "orders": OrderSerializer(manager_orders, many=True).data
        })

    return Response(data)




# ---------------- CUSTOMER MENU (QR) ----------------

@api_view(['GET'])
@permission_classes([AllowAny])
@authentication_classes([])
def customer_menu(request, restaurant_id, table_id):
    try:
        restaurant = Restaurant.objects.get(id=restaurant_id)
        ensure_restaurant_is_active(restaurant)

        table = Table.objects.get(id=table_id, restaurant=restaurant)
        categories = MenuCategory.objects.filter(restaurant=restaurant)

        menu_data = []
        for category in categories:
            items = MenuItemSerializer(
                category.items.filter(available=True),
                many=True
            ).data
            menu_data.append({
                'id': str(category.id),
                'name': category.name,
                'image': category.image,
                'items': items
            })

        return Response({
            'restaurant': RestaurantSerializer(restaurant).data,
            'table': TableSerializer(table).data,
            'menu': menu_data
        })

    except (Restaurant.DoesNotExist, Table.DoesNotExist):
        return Response({'error': 'Restaurant or Table not found'}, status=status.HTTP_404_NOT_FOUND)

# ---------------- BILLING ----------------

class BillViewSet(viewsets.ModelViewSet):
    queryset = Bill.objects.all()
    serializer_class = BillSerializer

    def get_queryset(self):
        queryset = Bill.objects.all()
        restaurant_id = self.request.query_params.get('restaurant_id')
        if restaurant_id:
            queryset = queryset.filter(session__table__restaurant_id=restaurant_id)
        return queryset.order_by('-created_at')

    @action(detail=False, methods=['post'])
    def generate(self, request):
        session_id = request.data.get('session_id')
        tax_rate = float(request.data.get('tax_rate', 0))

        try:
            session = OrderSession.objects.get(id=session_id)

            # 🔥 FIX: Check if user is authenticated and has role
            # If not authenticated (customer), allow them to generate bill for their own session
            if request.user.is_authenticated:
                # Staff permission check
                if request.user.role not in ['owner', 'restaurant_manager']:
                    raise PermissionDenied("You are not allowed to generate bills")
                
                # Staff can generate bill for any active session
                if session.status != 'active':
                    return Response(
                        {"error": "Bill already generated or session closed"},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            else:
                # Anonymous user (customer) - allow them to generate bill for their session
                # No role check needed, but we should verify the session belongs to their table
                if session.status != 'active':
                    return Response(
                        {"error": "Bill already generated or session closed"},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                # Optional: Verify session belongs to the table from request
                # This prevents customers from generating bills for other tables

            # Get restaurant from session
            if session.table:
                restaurant = session.table.restaurant
            elif session.table_group:
                restaurant = session.table_group.restaurant
            else:
                return Response(
                    {"error": "Session is not linked to table or group"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Check restaurant is active
            if request.user.is_authenticated and hasattr(request.user, 'role') and request.user.role == 'owner':
                ensure_restaurant_is_active(restaurant)

            # Prevent duplicate bill
            if hasattr(session, 'bill'):
                return Response(
                    {"error": "Bill already exists for this session"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Calculate totals
            orders = session.orders.all()
            subtotal = sum(order.total for order in orders)
            tax = subtotal * (Decimal(str(tax_rate)) / Decimal('100'))
            total = subtotal + tax

            bill = Bill.objects.create(
                session=session,
                subtotal=subtotal,
                tax=tax,
                total=total,
                generated_by=request.user if request.user.is_authenticated else None
            )

            # Update session status
            session.status = 'bill_generated'
            session.save()

            return Response(
                BillSerializer(bill).data,
                status=status.HTTP_201_CREATED
            )

        except OrderSession.DoesNotExist:
            return Response(
                {'error': 'Session not found'},
                status=status.HTTP_404_NOT_FOUND
            )
      
    @action(detail=True, methods=['patch'], url_path='update_payment')
    def mark_paid(self, request, pk=None):
        bill = self.get_object()
        user = request.user

        # 🔒 Only owner allowed
        if user.role != 'owner':
            raise PermissionDenied("Only owner can mark bill as paid")

        session = bill.session
        restaurant = session.table.restaurant
        ensure_restaurant_is_active(restaurant)

        # Prevent double payment
        if bill.payment_status == 'paid':
            return Response(
                {"error": "Bill already marked as paid"},
                status=status.HTTP_400_BAD_REQUEST
            )

        payment_mode = request.data.get('payment_mode')

        if payment_mode not in ['cash', 'upi', 'card']:
            return Response(
                {"error": "Invalid payment mode"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Update bill
        bill.payment_status = 'paid'
        bill.payment_mode = payment_mode
        bill.paid_at = timezone.now()
        bill.save()

        # Update session
        session.status = 'paid'
        session.end_time = timezone.now()
        session.save()

        # Free table
        # 🔁 Free tables depending on session type

        if session.table:
            session.table.status = 'available'
            session.table.save()

        elif session.table_group:
            group = session.table_group

            # Free all tables in group
            group.tables.update(status='available')

            # Deactivate group
            group.is_active = False
            group.save()


        return Response({"status": "paid"})


# Add to views.py - THIS IS A SEPARATE FUNCTION, NOT INSIDE THE CLASS
@api_view(['POST'])
@permission_classes([AllowAny])
@authentication_classes([])
def customer_generate_bill(request):
    """
    Customer can generate bill for their own session.
    No authentication required - they just need the session_id.
    """
    session_id = request.data.get('session_id')
    tax_rate = float(request.data.get('tax_rate', 5))

    try:
        session = OrderSession.objects.get(id=session_id)

        if session.status != 'active':
            return Response(
                {"error": "Bill already generated or session closed"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
          _ = session.bill
          return Response({"error": "Bill already exists"}, status=400)
        except Bill.DoesNotExist:
         pass

        # Calculate totals
        orders = session.orders.all()
        subtotal = sum(order.total for order in orders)
        tax = subtotal * (Decimal(str(tax_rate)) / Decimal('100'))
        total = subtotal + tax

        bill = Bill.objects.create(
            session=session,
            subtotal=subtotal,
            tax=tax,
            total=total,
            generated_by=None  # No user for customer-generated bills
        )

        session.status = 'bill_generated'
        session.save()

        return Response(
            BillSerializer(bill).data,
            status=status.HTTP_201_CREATED
        )

    except OrderSession.DoesNotExist:
        return Response(
            {'error': 'Session not found'},
            status=status.HTTP_404_NOT_FOUND
        )


# ---------------- USERS ----------------

from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied
from rest_framework.response import Response
from rest_framework import status

class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer

    def get_serializer_class(self):
        if self.action == 'create':
            return UserCreateSerializer
        return UserSerializer

    # ================= GET QUERYSET =================
    def get_queryset(self):
        user = self.request.user

        # 🔴 Super Admin → see all users
        if user.role == 'super_admin':
            return User.objects.all()

        # 🔵 Owner → see only his restaurant staff
        if user.role == 'owner':
            return User.objects.filter(
                restaurant=user.restaurant
            )

        # 🔵 Manager / Kitchen → cannot view users
        return User.objects.none()

    # ================= CREATE STAFF =================
    def perform_create(self, serializer):
        user = self.request.user
        role = self.request.data.get('role')

        # 🔐 Only owner can create staff
        if user.role != 'owner':
            raise PermissionDenied("Only owner can create staff accounts")

        # 🔐 Only allow specific roles
        if role not in ['kitchen_staff', 'restaurant_manager']:
            raise PermissionDenied("Invalid role selection")

        serializer.save(
            restaurant=user.restaurant
        )

    # ================= GET MANAGERS ONLY =================
    @action(detail=False, methods=['get'], url_path='managers')
    def get_managers(self, request):
      user = request.user
      if user.role != 'owner':
        raise PermissionDenied("Only owner can view managers")
      managers = User.objects.filter(
        restaurant=user.restaurant,
        role='restaurant_manager'
      )
      serializer = self.get_serializer(managers, many=True)
      return Response(serializer.data)

    # ✅ ADD THIS - Get kitchen staff only
    @action(detail=False, methods=['get'], url_path='kitchen-staff')
    def get_kitchen_staff(self, request):
      user = request.user
      if user.role != 'owner':
        raise PermissionDenied("Only owner can view kitchen staff")
      staff = User.objects.filter(
        restaurant=user.restaurant,
        role='kitchen_staff'
      )
      serializer = self.get_serializer(staff, many=True)
      return Response(serializer.data)

    # ================= TOGGLE STATUS =================
    @action(detail=True, methods=['patch'], url_path='toggle-status')
    def toggle_status(self, request, pk=None):
      user = request.user
      manager = self.get_object()
      if user.role != 'owner':
        raise PermissionDenied("Only owner can change staff status")
      # ✅ Allow both roles
      if manager.role not in ['restaurant_manager', 'kitchen_staff']:
        return Response({"error": "Invalid user"}, status=status.HTTP_400_BAD_REQUEST)
      manager.is_active = not manager.is_active
      manager.save()
      return Response({"message": "Staff status updated", "is_active": manager.is_active})

    # ================= RESET PASSWORD =================
    @action(detail=True, methods=['patch'], url_path='reset-password')
    def reset_password(self, request, pk=None):
      user = request.user
      manager = self.get_object()
      if user.role != 'owner':
        raise PermissionDenied("Only owner can reset password")
      # ✅ Allow both roles
      if manager.role not in ['restaurant_manager', 'kitchen_staff']:
        return Response({"error": "Invalid user"}, status=status.HTTP_400_BAD_REQUEST)
      new_password = request.data.get('password')
      if not new_password:
        return Response({"error": "Password required"}, status=status.HTTP_400_BAD_REQUEST)
      manager.set_password(new_password)
      manager.save()
      return Response({"message": "Password reset successful"})

    # ================= DELETE =================
    def destroy(self, request, *args, **kwargs):
      user = request.user
      manager = self.get_object()
      if user.role != 'owner':
        raise PermissionDenied("Only owner can delete staff")
      # ✅ Allow both roles
      if manager.role not in ['restaurant_manager', 'kitchen_staff']:
        raise PermissionDenied("You can only delete staff accounts")
      return super().destroy(request, *args, **kwargs)




# ---------------- HALLS ----------------

from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.exceptions import PermissionDenied
from .models import Hall, Table
from .serializers import HallSerializer

from rest_framework.permissions import IsAuthenticated

class HallViewSet(viewsets.ModelViewSet):
    serializer_class = HallSerializer
    queryset = Hall.objects.all()
    permission_classes = [IsAuthenticated]  # ✅ FIX: force auth

    def get_queryset(self):
        user = self.request.user
        restaurant_id = self.request.query_params.get('restaurant_id')

        # 🔒 BLOCK OWNER IF RESTAURANT INACTIVE
        if hasattr(user, 'role') and user.role == 'owner' and user.restaurant:
            ensure_restaurant_is_active(user.restaurant)
            return Hall.objects.filter(restaurant=user.restaurant)

        # ✅ SUPER ADMIN
        if hasattr(user, 'role') and user.role == 'super_admin':
            qs = Hall.objects.all()
            if restaurant_id:
                qs = qs.filter(restaurant_id=restaurant_id)
            return qs

        return Hall.objects.none()

    def perform_create(self, serializer):
        user = self.request.user

        # 🔒 BLOCK OWNER IF RESTAURANT INACTIVE
        if hasattr(user, 'role') and user.role == 'owner' and user.restaurant:
            ensure_restaurant_is_active(user.restaurant)
            serializer.save(restaurant=user.restaurant)
            return

        # ✅ SUPER ADMIN
        if hasattr(user, 'role') and user.role == 'super_admin':
            serializer.save()
            return

        raise PermissionDenied("You are not allowed to create halls")

    def destroy(self, request, *args, **kwargs):
        user = request.user
        hall = self.get_object()

        # 🔒 OWNER RULES
        if hasattr(user, 'role') and user.role == 'owner':
            if not user.restaurant:
                raise PermissionDenied("Owner has no restaurant")

            ensure_restaurant_is_active(user.restaurant)

            if hall.restaurant != user.restaurant:
                raise PermissionDenied("You cannot delete this hall")

        # 🧮 Count tables before delete (for UI message)
        tables_count = hall.tables.count()

        # ❌ DELETE HALL (tables deleted automatically via CASCADE)
        hall.delete()

        return Response(
            {
                'message': 'Hall deleted successfully',
                'tables_deleted': tables_count
            },
            status=status.HTTP_200_OK
        )



#----------------For menu items  --------#
from .models import MenuItemOption
from .serializers import (
    MenuItemOptionSerializer,
    MenuItemOptionWriteSerializer,
)
from .utils import ensure_restaurant_is_active, send_restaurant_status_email



class MenuItemOptionViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = MenuItemOption.objects.select_related(
        'item',
        'item__category',
        'item__category__restaurant'
    )

    def get_serializer_class(self):
        if self.request.method in ['POST', 'PUT', 'PATCH']:
            return MenuItemOptionWriteSerializer
        return MenuItemOptionSerializer

    def _check_owner_permissions(self, option):
        user = self.request.user

        # Super admin bypass
        if user.role == 'super_admin':
            return

        if user.role != 'owner':
            raise PermissionDenied("You do not have permission to manage item options")

        ensure_restaurant_is_active(user.restaurant)

        if option.item.category.restaurant != user.restaurant:
            raise PermissionDenied("You cannot manage options for this item")

    def perform_create(self, serializer):
        option = serializer.save()
        self._check_owner_permissions(option)

    def perform_update(self, serializer):
        option = serializer.instance
        self._check_owner_permissions(option)
        serializer.save()

    def perform_destroy(self, instance):
        self._check_owner_permissions(instance)
        instance.delete()


#-------Parcel views-------------------------------#
from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.exceptions import PermissionDenied
from django.shortcuts import get_object_or_404
from django.utils.timezone import now

from .models import (
    ParcelOrder,
    ParcelOrderItem,
    ParcelBill,
    Restaurant
)
from .serializers import (
    ParcelOrderSerializer,
    ParcelOrderCreateSerializer,
    ParcelBillSerializer
)
from .utils import ensure_restaurant_is_active

class ParcelOrderViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = ParcelOrderSerializer

    def get_queryset(self):
        user = self.request.user

        # 🔴 Super Admin → see all active parcels
        if user.role == 'super_admin':
            return ParcelOrder.objects.exclude(
                status__in=['completed', 'cancelled']
            )

        # 🔵 Owner / Staff → only their restaurant's active parcels
        if user.restaurant:
            return ParcelOrder.objects.filter(
                restaurant=user.restaurant
            ).exclude(
                status__in=['cancelled']
            ).exclude(
                status='completed',
                kitchen_completed=True
            )

        return ParcelOrder.objects.none()


    # ✅ OWNER: Accept parcel (send to kitchen)
    @action(detail=True, methods=['patch'])
    def accept(self, request, pk=None):
        parcel = self.get_object()

        if request.user.role != 'owner':
            raise PermissionDenied("Only owner can accept parcel orders")

        ensure_restaurant_is_active(parcel.restaurant)

        if parcel.status != 'pending':
            return Response(
                {"detail": "Only pending parcels can be accepted"},
                status=status.HTTP_400_BAD_REQUEST
            )

        parcel.status = 'accepted'
        parcel.accepted_at = timezone.now()
        parcel.save()

        return Response({"status": "accepted"})

    # ❌ OWNER: Cancel parcel (before preparation)
    @action(detail=True, methods=['patch'])
    def cancel(self, request, pk=None):
        parcel = self.get_object()

        if request.user.role != 'owner':
            raise PermissionDenied("Only owner can cancel parcel orders")

        ensure_restaurant_is_active(parcel.restaurant)

        if parcel.status == 'completed':
            return Response(
                {"detail": "Completed parcel cannot be cancelled"},
                status=status.HTTP_400_BAD_REQUEST
            )

        parcel.status = 'cancelled'
        parcel.save()

        return Response({"status": "cancelled"})

    # 💵 OWNER: Generate parcel bill
    @action(detail=True, methods=['post'])
    def generate_bill(self, request, pk=None):
        parcel = self.get_object()
        print(f"DEBUG >>> parcel.id={parcel.id} | parcel.status={parcel.status}")
        try:
         existing_bill = parcel.bill
         print(f"DEBUG >>> bill already exists: {existing_bill.id}")
        except ParcelBill.DoesNotExist:
         print(f"DEBUG >>> no bill exists")

        if request.user.role != 'owner':
            raise PermissionDenied("Only owner can generate parcel bill")

        ensure_restaurant_is_active(parcel.restaurant)

        if parcel.status != 'accepted':
            return Response(
                {"detail": "Parcel must be accepted before billing"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # ✅ CORRECT — actually try to access it
        try:
          _ = parcel.bill
          return Response(
        {"detail": "Parcel bill already exists"},
        status=status.HTTP_400_BAD_REQUEST
        )
        except ParcelBill.DoesNotExist:
          pass

        tax = Decimal(str(request.data.get('tax', 0)))
        payment_status = request.data.get('payment_status', 'pending')

        if payment_status not in ['paid', 'pending']:
            payment_status = 'pending'

        bill = ParcelBill.objects.create(
            parcel_order=parcel,
            subtotal=parcel.total,
            tax=tax,
            total=parcel.total + tax,
            payment_status=payment_status,
            paid_at=timezone.now() if payment_status == 'paid' else None
        )

        return Response(
            ParcelBillSerializer(bill).data,
            status=status.HTTP_201_CREATED
        )

    # ✅ OWNER: Mark parcel bill as paid
    @action(detail=True, methods=['patch'])
    def mark_paid(self, request, pk=None):
        parcel = self.get_object()

        if request.user.role != 'owner':
            raise PermissionDenied("Only owner can mark parcel as paid")

        ensure_restaurant_is_active(parcel.restaurant)

        payment_mode = request.data.get('payment_mode')        
        if payment_mode not in ['cash', 'upi', 'card']:        
          return Response(                                    
            {"error": "Invalid payment mode"},             
            status=status.HTTP_400_BAD_REQUEST                     )   

        try:
            bill = parcel.bill
        except ParcelBill.DoesNotExist:
            return Response(
                {"error": "No bill found for this parcel"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if bill.payment_status == 'paid':
            return Response(
                {"error": "Bill already marked as paid"},
                status=status.HTTP_400_BAD_REQUEST
            )

        bill.payment_status = 'paid'
        bill.payment_mode = payment_mode
        bill.paid_at = timezone.now()
        bill.save()

        return Response({"status": "paid"})
    
    # ✅ KITCHEN: Mark parcel as kitchen-completed
    @action(detail=True, methods=['patch'], url_path='complete-kitchen')
    def complete_kitchen(self, request, pk=None):
        parcel = self.get_object()

        if request.user.role not in ['owner', 'kitchen_staff']:
            raise PermissionDenied("Only kitchen staff or owner can complete parcel")

        if parcel.kitchen_completed:
            return Response(
                {"detail": "Parcel already marked as kitchen-completed"},
                status=status.HTTP_400_BAD_REQUEST
            )

        parcel.kitchen_completed = True
        parcel.kitchen_completed_at = timezone.now()
        parcel.save()

        # Broadcast to owner dashboard via WebSocket
        try:
            channel_layer = get_channel_layer()
            if channel_layer:
                async_to_sync(channel_layer.group_send)(
                    f'kitchen_{parcel.restaurant.id}',
                    {
                        'type': 'parcel_status_update',
                        'message': 'parcel_kitchen_completed',
                        'parcel_id': str(parcel.id),
                    }
                )
        except Exception as e:
            print(f"⚠️ WebSocket broadcast failed: {e}")

        return Response({"status": "kitchen_completed"})


    

@api_view(['GET'])
@permission_classes([AllowAny])
@authentication_classes([])
def parcel_menu(request, restaurant_id):
    restaurant = get_object_or_404(Restaurant, id=restaurant_id)
    ensure_restaurant_is_active(restaurant)

    if not restaurant.is_open:
        return Response(
            {'detail': 'Restaurant is currently closed', 'is_open': False},
            status=status.HTTP_403_FORBIDDEN
        )

    categories = MenuCategory.objects.filter(
        restaurant=restaurant
    ).prefetch_related('items__options')

    response = []

    for category in categories:
        items = category.items.filter(available=True)

        response.append({
            "id": category.id,
            "name": category.name,
            "restaurant_name": restaurant.name,
            "items": [
                {
                    "id": item.id,
                    "name": item.name,
                    "price": item.price,
                    "half_price": item.half_price,
                    "allow_half": item.allow_half,
                    "options": [
                        {
                            "id": opt.id,
                            "name": opt.name,
                            "extra_price": opt.extra_price
                        }
                        for opt in item.options.filter(is_active=True)
                    ]
                }
                for item in items
            ]
        })

    return Response(response)


@api_view(['POST'])
@permission_classes([AllowAny])
@authentication_classes([])
def create_parcel_order(request, restaurant_id):
    restaurant = get_object_or_404(Restaurant, id=restaurant_id)
    ensure_restaurant_is_active(restaurant)

    if not restaurant.is_open:
        return Response(
            {'detail': 'Restaurant is currently closed', 'is_open': False},
            status=status.HTTP_403_FORBIDDEN
        )

    serializer = ParcelOrderCreateSerializer(
        data=request.data,
        context={
            'request': request,
            'restaurant': restaurant
        }
    )

    serializer.is_valid(raise_exception=True)
    order = serializer.save()
    # In create_parcel_order function, after serializer.save() returns the order:

    # Broadcast to kitchen for real-time update
    channel_layer = get_channel_layer()
    async_to_sync(channel_layer.group_send)(
        f'kitchen_{restaurant.id}',
      {
        'type': 'kitchen_order_update',
        'order': {
            'id': str(order.id),
            'type': 'parcel',
            'customer_name': order.customer_name,
            'customer_phone': order.customer_phone,
            'total': str(order.total),
            'status': order.status,
            'created_at': order.created_at.isoformat(),
            'restaurant_name': restaurant.name,
            'hall_name': 'Parcel',
            'items': [
                {
                    'id': str(item.id),
                    'name': item.menu_item.name,
                    'quantity': item.quantity,
                    'display_quantity': float(item.quantity),
                    'price': str(item.price),
                    'completed': False,
                    'selected_options': [  # ✅ ADD THIS
                        {
                            'id': str(opt.id),
                            'name': opt.name,
                            'extra_price': str(opt.extra_price)
                        }
                        for opt in item.selected_options.all()
                    ]
                    
                }
                for item in order.items.all()
            ]
        },
        'message': 'New parcel order received'
    }
)

    return Response(
        ParcelOrderSerializer(order).data,
        status=status.HTTP_201_CREATED
    )


# ================= PARCEL  Bill =================
from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from .models import ParcelBill
from .serializers import ParcelBillSerializer


class ParcelBillViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = ParcelBillSerializer

    def get_queryset(self):
        user = self.request.user

        # 🔴 Super Admin → all parcel bills
        if user.role == 'super_admin':
            return ParcelBill.objects.all()

        # 🔵 Owner / staff → only their restaurant
        if user.restaurant:
            ensure_restaurant_is_active(user.restaurant)
            return ParcelBill.objects.filter(
                parcel_order__restaurant=user.restaurant
            )

        return ParcelBill.objects.none()
    

#---for qr and bill preview----------#
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from .models import RestaurantPrintSettings
from .serializers import RestaurantPrintSettingsSerializer
from .utils import ensure_restaurant_is_active


class RestaurantPrintSettingsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user

        # 🔴 Super Admin → not allowed here (optional rule)
        if user.role == 'super_admin':
            return Response(
                {"detail": "Super admin cannot access print settings"},
                status=403
            )

        # 🔵 Owner / staff
        if user.restaurant:
            ensure_restaurant_is_active(user.restaurant)

            settings, _ = RestaurantPrintSettings.objects.get_or_create(
                restaurant=user.restaurant
            )

            serializer = RestaurantPrintSettingsSerializer(settings)
            return Response(serializer.data)

        return Response(
            {"detail": "User is not linked to a restaurant"},
            status=400
        )

    def put(self, request):
        user = request.user

        # 🔴 Super Admin → not allowed
        if user.role == 'super_admin':
            return Response(
                {"detail": "Super admin cannot update print settings"},
                status=403
            )

        # 🔵 Owner / staff
        if user.restaurant:
            ensure_restaurant_is_active(user.restaurant)

            settings, _ = RestaurantPrintSettings.objects.get_or_create(
                restaurant=user.restaurant
            )

            serializer = RestaurantPrintSettingsSerializer(
                settings,
                data=request.data,
                partial=True
            )
            serializer.is_valid(raise_exception=True)
            serializer.save()

            return Response(serializer.data)

        return Response(
            {"detail": "User is not linked to a restaurant"},
            status=400
        )


#------------------For Excel-----------------#
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def today_sales_excel(request):
    user = request.user

    # 🔐 Only owner / staff
    if not user.restaurant:
        return Response({'error': 'No restaurant'}, status=403)

    restaurant = user.restaurant

    # 🚫 restaurant inactive
    ensure_restaurant_is_active(restaurant)

    today = timezone.now().date()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Today Sales'

    # 🟧 TITLE
    ws.merge_cells('A1:E1')
    ws['A1'] = f"Today's Sales Report - {today.strftime('%d %b %Y')}"

    # 🟨 HEADERS
    ws.append([
        'Date',
        'Order Type',
        'Customer Name',
        'Phone',
        'Total Amount'
    ])

    total_sum = 0

    # 🧾 TABLE BILLS
    table_bills = Bill.objects.filter(
        session__table__restaurant_id=restaurant.id,
        created_at__date=today
    )

    for bill in table_bills:
        ws.append([
            bill.created_at.strftime('%Y-%m-%d'),
            'Table',
            bill.session.customer_name or 'Guest',
            bill.session.customer_phone or '',
            float(bill.total)
        ])
        total_sum += bill.total

    # 📦 PARCEL BILLS
    parcel_bills = ParcelBill.objects.filter(
        parcel_order__restaurant=restaurant,
        created_at__date=today
    )

    for bill in parcel_bills:
        ws.append([
            bill.created_at.strftime('%Y-%m-%d'),
            'Parcel',
            bill.parcel_order.customer_name,
            bill.parcel_order.customer_phone,
            float(bill.total)
        ])
        total_sum += bill.total

    # 🟩 TOTAL ROW
    ws.append([])
    ws.append(['', '', '', 'TOTAL', float(total_sum)])

    # 📤 RESPONSE
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename=today-sales-{today}.xlsx'
    )

    wb.save(response)
    return response


#---New Close rule view-----------------------#
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from django.shortcuts import get_object_or_404

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def close_restaurant_day(request, restaurant_id):
    user = request.user

    # 🔐 Permission
    if user.role not in ['owner', 'super_admin']:
        return Response({'error': 'Permission denied'}, status=403)

    restaurant = get_object_or_404(Restaurant, id=restaurant_id)

    if user.role == 'owner' and restaurant.owner != user:
        return Response({'error': 'Permission denied'}, status=403)

    # ===============================
    # 🔥 TABLE FLOW (STRICT ORDER)
    # ===============================

    # Get all sessions of this restaurant
    sessions = OrderSession.objects.filter(
        table__restaurant=restaurant
    ) | OrderSession.objects.filter(
        table_group__restaurant=restaurant
    )

    # 1️⃣ Delete Bills
    Bill.objects.filter(session__in=sessions).delete()

    # 2️⃣ Delete OrderItems
    OrderItem.objects.filter(order__session__in=sessions).delete()

    # 3️⃣ Delete Orders
    Order.objects.filter(session__in=sessions).delete()

    # 4️⃣ Delete Sessions
    sessions.delete()


    # 5️⃣ Reset tables (IMPORTANT)
    Table.objects.filter(
        restaurant=restaurant
    ).update(status='available')

    # ===============================
    # 📦 PARCEL FLOW (STRICT ORDER)
    # ===============================

    # 6️⃣ Delete Parcel Bills
    ParcelBill.objects.filter(
        parcel_order__restaurant=restaurant
    ).delete()

    # 7️⃣ Delete Parcel Order Items (if exists)
    ParcelOrderItem.objects.filter(
        parcel_order__restaurant=restaurant
    ).delete()

    # 8️⃣ Delete Parcel Orders
    ParcelOrder.objects.filter(
        restaurant=restaurant
    ).delete()

    # ===============================
    # 🔒 CLOSE RESTAURANT
    # ===============================
    restaurant.is_open = False
    restaurant.save(update_fields=['is_open'])

    return Response(
        {'message': 'Restaurant closed & ALL operational data cleared'},
        status=200
    )


class OrderSessionViewSet(viewsets.ModelViewSet):
    queryset = OrderSession.objects.all()
    serializer_class = OrderSessionSerializer
    
    def get_queryset(self):
        queryset = OrderSession.objects.all()
        
        # Filter by query params
        restaurant_id = self.request.query_params.get('restaurant')
        table_id = self.request.query_params.get('table')
        status = self.request.query_params.get('status')
        
        if restaurant_id:
            queryset = queryset.filter(
                Q(table__restaurant_id=restaurant_id) | 
                Q(table_group__restaurant_id=restaurant_id)
            )
        if table_id:
            queryset = queryset.filter(table_id=table_id)
        if status:
            queryset = queryset.filter(status=status)
            
        return queryset.select_related('table', 'table_group').prefetch_related('orders', 'orders__items')
    
    # Allow customers to view their own session
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return []  # Public access for customers
        return [IsAuthenticated()]  # Staff only for modifications
    


# ================= KITCHEN DASHBOARD VIEWS =================#

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Q
from .models import Order, OrderSession, ParcelOrder


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def kitchen_orders(request):
    """
    Get all orders for kitchen dashboard.
    Returns categorized orders: first_orders, live_orders, parcel_orders
    """
    user = request.user
    
    if user.role not in ['owner', 'kitchen_staff']:
        raise PermissionDenied("Only owner or kitchen staff can access kitchen")
    
    restaurant = user.restaurant
    if not restaurant:
        return Response({'error': 'No restaurant assigned'}, status=400)
    
    ensure_restaurant_is_active(restaurant)
    
    # Get active sessions with orders - INCLUDE hall data
    active_sessions = OrderSession.objects.filter(
        Q(table__restaurant=restaurant) | Q(table_group__restaurant=restaurant),
        status='active'
    ).select_related(
        'table', 
        'table__hall',  # ✅ ADD: Fetch hall with table
        'table__restaurant',  # ✅ ADD: Fetch restaurant
        'table_group'
    ).prefetch_related('orders', 'orders__items', 'orders__items__menu_item')
    
    first_orders = []
    live_orders = []
    
    for session in active_sessions:

        # Get all orders sorted
        all_orders = list(session.orders.all().order_by('created_at'))

        # true first order of session (even if completed)
        true_first_order_id = None
        if all_orders:
            true_first_order_id = all_orders[0].id

        # active orders (not completed)
        active_orders = [o for o in all_orders if o.status != "completed"]

        if not active_orders:
            continue

        # ✅ ADD: Get hall name from table
        hall_name = None
        if session.table and session.table.hall:
            hall_name = session.table.hall.name
        elif session.table_group:
            # For combined tables, get hall from first table in group
            first_table = session.table_group.tables.first()
            if first_table and first_table.hall:
                hall_name = first_table.hall.name
        
        for order in active_orders:

            order_data = {
                'id': str(order.id),
                'session_id': str(session.id),
                'status': order.status,
                'order_source': order.order_source,
                'total': str(order.total),
                'special_instructions': order.special_instructions,
                'created_at': order.created_at.isoformat(),
                'table_number': session.table.table_number if session.table else None,
                'table_group': str(session.table_group.id) if session.table_group else None,
                'customer_name': session.customer_name or 'Guest',
                'restaurant_name': restaurant.name,  # ✅ ADD: Restaurant name
                'hall_name': hall_name or 'General',  # ✅ ADD: Hall name
                'is_first_order': order.id == true_first_order_id,
                'items': []
            }

            if session.table_group:
                order_data['group_tables'] = [
                    t.table_number for t in session.table_group.tables.all()
                ]

            for item in order.items.all():
                # ✅ FIX: Calculate display quantity for half plates
                display_quantity = float(item.quantity)
                if item.is_half:
                    display_quantity += 0.5
                
                selected_options = [
                    {
                        'id': str(opt.id),
                        'name': opt.name,
                        'extra_price': str(opt.extra_price)
                    }
                    for opt in item.selected_options.all()
                ]
                
                order_data['items'].append({
                    'id': str(item.id),
                    'name': item.menu_item.name,
                    'quantity': item.quantity,  # Keep original for reference
                    'display_quantity': display_quantity,  # ✅ ADD: Calculated quantity (0.5, 1, 1.5, etc.)
                    'is_half': item.is_half,  # Keep for reference if needed
                    'price': str(item.price),
                    'completed': False,
                    'selected_options': selected_options
                })

            if order.id == true_first_order_id:
                first_orders.append(order_data)
            else:
                live_orders.append(order_data)
    
    # Get accepted parcel orders — only show if kitchen has NOT completed them
    parcel_orders = ParcelOrder.objects.filter(
        restaurant=restaurant,
        status='accepted',
        kitchen_completed=False
    ).prefetch_related('items', 'items__menu_item')
    
    parcel_data = []
    for parcel in parcel_orders:
        parcel_item = {
            'id': str(parcel.id),
            'type': 'parcel',
            'customer_name': parcel.customer_name,
            'customer_phone': parcel.customer_phone,
            'total': str(parcel.total),
            'status': parcel.status,
            'created_at': parcel.created_at.isoformat(),
            'restaurant_name': restaurant.name,  # ✅ ADD: Restaurant name for parcels too
            'hall_name': 'Parcel',  # ✅ ADD: Label for parcel orders
            'items': []
            
        }
        
        for item in parcel.items.all():
            # Parcel items don't have is_half, so quantity is as-is
            parcel_item['items'].append({
                'id': str(item.id),
                'name': item.menu_item.name,
                'quantity': item.quantity,
                'display_quantity': float(item.quantity),  # ✅ ADD: Consistent with table orders
                'price': str(item.price),
                'completed': False,
                'selected_options': [
            {
                'id': str(opt.id),
                'name': opt.name,
                'extra_price': str(opt.extra_price)
            }
            for opt in item.selected_options.all()
        ],
            })
        
        parcel_data.append(parcel_item)
    
    return Response({
        'restaurant_name': restaurant.name,
    'is_open': restaurant.is_open,
    'first_orders': first_orders,
    'live_orders': live_orders,
    'parcel_orders': parcel_data
    })


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def update_order_status(request, order_id):
    """Update order status (pending -> preparing -> ready -> completed)"""
    user = request.user
    
    if user.role not in ['owner', 'kitchen_staff']:
        raise PermissionDenied("Only owner or kitchen staff can update orders")
    
    try:
        order = Order.objects.get(id=order_id)
    except Order.DoesNotExist:
        return Response({'error': 'Order not found'}, status=404)
    
    new_status = request.data.get('status')
    if new_status not in ['pending', 'preparing', 'ready', 'completed']:
        return Response({'error': 'Invalid status'}, status=400)
    
    order.status = new_status
    order.save()
    
    return Response({
        'id': str(order.id),
        'status': order.status,
        'message': f'Order status updated to {new_status}'
    })


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def update_parcel_status(request, parcel_id):
    """Update parcel order status"""
    user = request.user
    
    if user.role not in ['owner', 'kitchen_staff']:
        raise PermissionDenied("Only owner or kitchen staff can update parcels")
    
    try:
        parcel = ParcelOrder.objects.get(id=parcel_id)
    except ParcelOrder.DoesNotExist:
        return Response({'error': 'Parcel order not found'}, status=404)
    
    new_status = request.data.get('status')
    valid_statuses = ['accepted', 'completed', 'cancelled']
    
    if new_status not in valid_statuses:
        return Response({'error': 'Invalid status'}, status=400)
    
    parcel.status = new_status
    parcel.save()
    
    return Response({
        'id': str(parcel.id),
        'status': parcel.status,
        'message': f'Parcel status updated to {new_status}'
    })


# ---------------- WEBSOCKET BROADCAST HELPERS ----------------


from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync


def broadcast_to_kitchen(order, restaurant_id, is_first_order=False):
    """
    Broadcast new order to kitchen dashboard via WebSocket.
    If WebSocket fails, order still saves successfully (kitchen will see it on next poll).
    """
    try:
        channel_layer = get_channel_layer()
        if not channel_layer:
            print("⚠️ No channel layer available, skipping WebSocket broadcast")
            return  # Silently skip - kitchen will see order on next poll
            
        # Determine order type
        order_type = 'first' if is_first_order else 'live'
        
        # Get session and related data safely
        session = order.session
        
        # Build order data (same structure as kitchen_orders API)
        order_data = {
            'id': str(order.id),
            'session_id': str(session.id),
            'status': order.status,
            'order_source': order.order_source,
            'total': str(order.total),
            'special_instructions': order.special_instructions,
            'created_at': order.created_at.isoformat(),
            'table_number': session.table.table_number if session.table else None,
            'table_group': str(session.table_group.id) if session.table_group else None,
            'customer_name': session.customer_name or 'Guest',
            'restaurant_name': session.table.restaurant.name if session.table else (
                session.table_group.restaurant.name if session.table_group else 'Unknown'
            ),
            'hall_name': (
                session.table.hall.name 
                if session.table and session.table.hall 
                else 'General'
            ),
            'is_first_order': is_first_order,
            'type': order_type,
            'items': []
        }

        # Add group tables if applicable
        if session.table_group:
            order_data['group_tables'] = [
                t.table_number for t in session.table_group.tables.all()
            ]

        # Build items with display_quantity
        for item in order.items.all():
            display_quantity = float(item.quantity)
            if item.is_half:
                display_quantity += 0.5

            selected_options = [
                    {
                        'id': str(opt.id),
                        'name': opt.name,
                        'extra_price': str(opt.extra_price)
                    }
                    for opt in item.selected_options.all()
                ]    
            
                
                
            order_data['items'].append({
                'id': str(item.id),
                'name': item.menu_item.name,
                'quantity': item.quantity,
                'display_quantity': display_quantity,
                'is_half': item.is_half,
                'price': str(item.price),
                'completed': False,
                'selected_options': selected_options
            })

        # Send to kitchen group
        async_to_sync(channel_layer.group_send)(
            f'kitchen_{restaurant_id}',
            {
                'type': 'kitchen_order_update',
                'order': order_data,
                'message': f'New {order_type} order received'
            }
        )
        
        print(f"📡 Broadcasted {order_type} order {order.id} to kitchen_{restaurant_id}")
        
    except Exception as e:
        # Log error but don't break order creation
        print(f"⚠️ WebSocket broadcast failed (order saved successfully): {e}")
        # Kitchen will see the order on next 30-second poll